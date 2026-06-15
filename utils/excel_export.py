"""openpyxl による Excel 出力（表示ラベル使用）。"""
import datetime
import io
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import holidays as holidays_lib
from utils.time_utils import SHIFT_LABEL, is_holiday

LABEL_COLORS = {
    "日":  "DDEEFF",
    "オ2": "EEFFDD",
    "ヤ1": "FFE0B2",
    "ヤ2": "FFCC80",
    "休":  "F5F5F5",
    "有":  "FCE4EC",
    "研修": "E8F5E9",
    "/イ": "FFF9C4",   # 薄黄色
}
SHIFT_HOURS = {"D": 7.5, "L": 7.5, "N1": 8.0, "N2": 8.0, "O": 0.0, "P": 0.0}

HEADER_FILL  = PatternFill("solid", fgColor="37474F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=9)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WEEKDAY_NAMES = "月火水木金土日"


def _cell_fill(label: str) -> PatternFill:
    return PatternFill("solid", fgColor=LABEL_COLORS.get(label, "FFFFFF"))


def _header(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    return c


def export_excel(
    schedule_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates: List[datetime.date],
    violations: List[Dict],
    hospital_holidays=(),
) -> bytes:
    wb = Workbook()

    # ── シート1: 勤務表 ────────────────────────────────────────
    ws = wb.active
    ws.title = "勤務表"

    sid_to_name = {row["id"]: row["name"] for _, row in staff_df.iterrows()}
    sid_to_target = {row["id"]: row["target_hours"] for _, row in staff_df.iterrows()}

    # ヘッダー行
    _header(ws, 1, 1, "氏名")
    ws.column_dimensions["A"].width = 14
    jp_hols = holidays_lib.Japan(years={d.year for d in dates})
    _hosp_hols_set = set(hospital_holidays)
    def _col_label(d):
        if d in _hosp_hols_set:
            return f"{d.month}/{d.day}\n病休"
        if d in jp_hols:
            return f"{d.month}/{d.day}\n祝"
        return f"{d.month}/{d.day}\n{WEEKDAY_NAMES[d.weekday()]}"

    for ci, d in enumerate(dates, start=2):
        label = _col_label(d)
        c = _header(ws, 1, ci, label)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = 4.5
    _header(ws, 1, len(dates) + 2, "夜勤回数")
    _header(ws, 1, len(dates) + 3, "勤務時間(h)")
    ws.row_dimensions[1].height = 30

    for ri, staff_id in enumerate(schedule_df.index, start=2):
        name_cell = ws.cell(ri, 1, sid_to_name.get(staff_id, str(staff_id)))
        name_cell.border = BORDER
        name_cell.font = Font(size=9)
        night_count = 0
        total_hours = 0.0
        for ci, d in enumerate(dates, start=2):
            code = schedule_df.loc[staff_id, d] if d in schedule_df.columns else "O"
            label = SHIFT_LABEL.get(code, code)
            c = ws.cell(ri, ci, label)
            c.fill = _cell_fill(label)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
            c.font = Font(size=9)
            if code == "N1":
                night_count += 1
            total_hours += SHIFT_HOURS.get(code, 0.0)
        ws.cell(ri, len(dates) + 2, night_count).border = BORDER
        ws.cell(ri, len(dates) + 2).alignment = Alignment(horizontal="center")
        ws.cell(ri, len(dates) + 3, round(total_hours, 1)).border = BORDER
        ws.cell(ri, len(dates) + 3).alignment = Alignment(horizontal="center")

    # 日別集計行
    ri_sum = len(schedule_df) + 2
    ws.cell(ri_sum, 1, "【日計】").font = Font(bold=True, size=9)
    for ci, d in enumerate(dates, start=2):
        day_col = schedule_df[d] if d in schedule_df.columns else pd.Series()
        summary = (
            f"日:{(day_col=='D').sum()} "
            f"オ2:{(day_col=='L').sum()}\n"
            f"ヤ1:{(day_col=='N1').sum()} "
            f"ヤ2:{(day_col=='N2').sum()}"
        )
        c = ws.cell(ri_sum, ci, summary)
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.font = Font(size=8)
        c.border = BORDER
    ws.row_dimensions[ri_sum].height = 30

    # ── シート2: スタッフ別集計 ────────────────────────────────
    ws2 = wb.create_sheet("スタッフ別集計")
    headers2 = ["氏名", "日", "オ2", "ヤ1", "ヤ2", "休", "有", "夜勤回数", "勤務時間(h)", "目標(h)", "差異(h)"]
    for ci, h in enumerate(headers2, 1):
        _header(ws2, 1, ci, h)
    ws2.column_dimensions["A"].width = 14

    for ri, staff_id in enumerate(schedule_df.index, start=2):
        row_codes = [schedule_df.loc[staff_id, d] for d in dates if d in schedule_df.columns]
        ws2.cell(ri, 1, sid_to_name.get(staff_id, str(staff_id))).font = Font(size=9)
        counts = {c: row_codes.count(c) for c in ["D", "L", "N1", "N2", "O", "P"]}
        for ci, code in enumerate(["D", "L", "N1", "N2", "O", "P"], 2):
            ws2.cell(ri, ci, counts[code]).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 8, counts["N1"]).alignment = Alignment(horizontal="center")
        hours = sum(SHIFT_HOURS.get(c, 0.0) for c in row_codes)
        target = sid_to_target.get(staff_id, 170.0)
        ws2.cell(ri, 9, round(hours, 1)).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 10, target).alignment = Alignment(horizontal="center")
        diff_cell = ws2.cell(ri, 11, round(hours - target, 1))
        diff_cell.alignment = Alignment(horizontal="center")
        if hours - target > 10:
            diff_cell.fill = PatternFill("solid", fgColor="FFCCBC")
        elif hours - target < -10:
            diff_cell.fill = PatternFill("solid", fgColor="BBDEFB")

    # ── シート3: 制約違反 ──────────────────────────────────────
    ws3 = wb.create_sheet("制約違反")
    for ci, h in enumerate(["種別", "日付", "スタッフ", "詳細"], 1):
        _header(ws3, 1, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = 26
    for ri, v in enumerate(violations, 2):
        sid = v.get("staff_id")
        name = sid_to_name.get(sid, "") if sid else ""
        ws3.cell(ri, 1, v.get("type", ""))
        ws3.cell(ri, 2, str(v.get("date", "")))
        ws3.cell(ri, 3, name)
        ws3.cell(ri, 4, v.get("detail", ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(schedule_df: pd.DataFrame, staff_df: pd.DataFrame,
               dates: List[datetime.date], hospital_holidays=()) -> bytes:
    """表示ラベルで CSV 出力。"""
    sid_to_name = {row["id"]: row["name"] for _, row in staff_df.iterrows()}
    out = schedule_df.copy()
    out.index = [sid_to_name.get(i, str(i)) for i in out.index]
    jp_hols2 = holidays_lib.Japan(years={d.year for d in dates})
    _hosp_set = set(hospital_holidays)
    def _csv_col(d):
        if d in _hosp_set:
            return f"{d.month}/{d.day}(病休)"
        if d in jp_hols2:
            return f"{d.month}/{d.day}(祝)"
        return f"{d.month}/{d.day}({WEEKDAY_NAMES[d.weekday()]})"
    out.columns = [_csv_col(d) for d in dates]
    # 内部コード → 表示ラベル
    out = out.map(lambda v: SHIFT_LABEL.get(v, v))
    return out.to_csv().encode("utf-8-sig")
