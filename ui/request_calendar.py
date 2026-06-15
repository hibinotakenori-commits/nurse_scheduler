"""スタッフごとの勤務希望カレンダー入力（月カレンダー方式 + テキスト一括入力 + Excelアップロード）。"""
import calendar as cal_module
import datetime
import io
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
import holidays as holidays_lib

from utils.time_utils import REQUEST_OPTIONS, REQUEST_TO_SHIFT
from utils.settings import save_requests

WEEKDAY_NAMES = "月火水木金土日"
_HOL_COLOR  = "#d32f2f"
_SAT_COLOR  = "#1565c0"
_NORM_COLOR = "#212121"
_GRAY_COLOR = "#bdbdbd"

# カレンダーセル内では短いラベルを使う
SHORT_OPTIONS = ["－", "日希", "夜希", "夜確", "深確", "研修", "委員", "認定", "休希", "有休"]
_FULL_TO_SHORT = dict(zip(REQUEST_OPTIONS, SHORT_OPTIONS))
_SHORT_TO_FULL = dict(zip(SHORT_OPTIONS, REQUEST_OPTIONS))

# 選択肢ごとのバッジ色
_BADGE_COLOR = {
    "日希": ("#1565c0", "#ffffff"),
    "夜希": ("#e65100", "#ffffff"),
    "夜確": ("#bf360c", "#ffffff"),   # 夜勤確定（ヤ1）= 深い赤オレンジ
    "深確": ("#4e342e", "#ffffff"),   # 深夜確定（ヤ2）= ダークブラウン
    "研修": ("#2e7d32", "#ffffff"),
    "委員": ("#f57f17", "#ffffff"),   # 委員会 = 濃い黄色
    "認定": ("#00695c", "#ffffff"),   # 認定看護師業務 = ティール
    "休希": ("#546e7a", "#ffffff"),
    "有休": ("#6a1b9a", "#ffffff"),
}

# ──────────────────────────────────────────────────────────────
# 共通：シフトキーワードテーブル（テキスト・Excel 両パーサーで使用）
# ──────────────────────────────────────────────────────────────

_SHIFT_KW_TABLE: List[Tuple[str, str, bool]] = [
    # ── 長いキーワードを先に（部分マッチ防止） ──
    ("準夜勤",  "N1", True),   # 準夜勤 → 夜勤入り
    ("深夜勤",  "N2", True),   # 深夜勤 → 夜勤明け
    ("準夜",    "N1", True),
    ("深夜",    "N2", True),
    ("夜勤確定","N1", True),
    ("夜勤希望","N1", False),
    ("夜勤",    "N1", False),
    ("日勤確定","D",  True),
    ("日勤希望","D",  False),
    ("日勤",    "D",  False),
    ("日/研",   "T",  True),   # 日勤＋研修を研修として記録
    ("研/日",   "T",  True),
    ("有休申請","P",  True),
    ("有給申請","P",  True),
    ("有休",    "P",  True),
    ("有給",    "P",  True),
    ("研修",    "T",  True),
    ("認定日",  "C",  True),   # 認定日 → 認定看護師業務
    ("認定",    "C",  True),   # 認定   → 認定看護師業務
    ("主任会",  "I",  True),   # 主任会議 → 委員会扱い
    ("委員会",  "I",  True),
    ("/イ",     "I",  True),
    ("委員",    "I",  True),
    ("休み確定","O",  True),
    ("休み希望","O",  False),
    ("お休み",  "O",  False),
    ("休暇",    "O",  False),
    ("休み",    "O",  False),
    ("遅出",    "L",  False),
    ("オ2",     "L",  False),
    ("準",      "N1", True),   # 準夜勤の略（「準」単独）
    ("深",      "N2", True),   # 深夜勤の略（「深」単独）
    # ショートラベル
    ("夜希",    "N1", False),
    ("日希",    "D",  False),
    ("休希",    "O",  False),
]

_LABEL_MAP: Dict[Tuple[str, bool], str] = {
    ("N1", False): "夜勤希望（準）",  ("N1", True): "夜勤確定（準）",
    ("N2", False): "深夜希望（深）",  ("N2", True): "深夜確定（深）",
    ("D",  False): "日勤希望",        ("D",  True): "日勤確定",
    ("O",  False): "休み希望",        ("O",  True): "休み確定",
    ("P",  True):  "有休申請",
    ("T",  True):  "研修",
    ("I",  True):  "委員会",
    ("C",  True):  "認定",
    ("L",  False): "遅出希望",
}

# ──────────────────────────────────────────────────────────────
# 自然言語パーサー
# ──────────────────────────────────────────────────────────────

# 自然言語パーサー用：_SHIFT_KW_TABLE を参照
_SHIFT_KEYWORDS = _SHIFT_KW_TABLE


def _find_staff(text: str, staff_df: pd.DataFrame) -> Optional[Tuple[int, str]]:
    """テキスト中のスタッフ名を検索。(staff_id, name) or None を返す。"""
    candidates = []
    for _, row in staff_df.iterrows():
        name: str = row["name"]
        name_nospace = name.replace(" ", "").replace("　", "")
        parts = name.split()
        last = parts[0] if parts else name
        for pattern in [name, name_nospace, last]:
            if pattern and pattern in text:
                candidates.append((len(pattern), row["id"], name))
                break
    if not candidates:
        return None
    candidates.sort(reverse=True)  # 長いマッチを優先
    _, sid, sname = candidates[0]
    return sid, sname


def _resolve_dates(
    day_nums: List[int],
    month_hint: Optional[int],
    dates: List[datetime.date],
) -> List[datetime.date]:
    """日番号リストをスケジュール期間内の date に解決する。"""
    date_by_md: Dict[Tuple[int, int], datetime.date] = {
        (d.month, d.day): d for d in dates
    }
    months_in_period = list(dict.fromkeys(d.month for d in dates))

    result = []
    for day in day_nums:
        if month_hint is not None:
            key = (month_hint, day)
            if key in date_by_md:
                result.append(date_by_md[key])
        else:
            # 月の指定なし → 期間内の全月で探す
            for m in months_in_period:
                key = (m, day)
                if key in date_by_md:
                    result.append(date_by_md[key])
    return result


def parse_request_text(
    text: str,
    staff_df: pd.DataFrame,
    dates: List[datetime.date],
) -> Tuple[List[Dict], List[str], List[str]]:
    """
    自然言語テキストを解析して希望リストを返す。

    Returns:
        parsed:   [{staff_id, staff_name, date, shift, is_fixed, label}, ...]
        warnings: 解釈できなかった部分などの警告
        errors:   致命的なエラー
    """
    parsed: List[Dict] = []
    warnings: List[str] = []
    errors: List[str] = []

    text = text.strip()
    if not text:
        return parsed, warnings, errors

    # ── スタッフ検索 ──
    found = _find_staff(text, staff_df)
    if found is None:
        errors.append("スタッフ名が見つかりませんでした。氏名（姓だけでも可）を含めてください。")
        return parsed, warnings, errors
    staff_id, staff_name = found

    # ── テキストを「、。,，\n」で分割 ──
    segments = re.split(r"[、。,，\n]+", text)

    date_by_md = {(d.month, d.day): d for d in dates}

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue

        # ── シフト判定 ──
        shift_code: Optional[str] = None
        is_fixed = False
        for kw, code, fixed in _SHIFT_KEYWORDS:
            if kw in seg:
                shift_code = code
                is_fixed = fixed
                break
        # 「確定」が単独で付いている場合（例:「日勤は確定」）
        if shift_code and not is_fixed and "確定" in seg:
            is_fixed = True

        if shift_code is None:
            continue  # このセグメントにシフト情報なし

        # ── 日付抽出 ──
        found_dates: List[datetime.date] = []

        # 1) 範囲: X日〜Y日 / X日からY日
        for m in re.finditer(r"(\d{1,2})日[〜～からーから]{1,3}(\d{1,2})日", seg):
            d1, d2 = int(m.group(1)), int(m.group(2))
            month_m = re.search(r"(\d{1,2})月", seg)
            mhint = int(month_m.group(1)) if month_m else None
            for day in range(d1, d2 + 1):
                found_dates += _resolve_dates([day], mhint, dates)

        # 2) M/D 形式（例: 7/24）
        for m in re.finditer(r"(\d{1,2})/(\d{1,2})", seg):
            month, day = int(m.group(1)), int(m.group(2))
            key = (month, day)
            if key in date_by_md:
                found_dates.append(date_by_md[key])

        # 3) X月Y日
        out_of_range_warned = False
        for m in re.finditer(r"(\d{1,2})月(\d{1,2})日", seg):
            month, day = int(m.group(1)), int(m.group(2))
            key = (month, day)
            if key in date_by_md:
                found_dates.append(date_by_md[key])
            else:
                warnings.append(f"{month}月{day}日はスケジュール期間外です")
                out_of_range_warned = True

        # 4) Y日（月の指定なし） — 上記でマッチした部分を除去してから処理
        seg_stripped = re.sub(r"\d{1,2}月\d{1,2}日", "", seg)
        seg_stripped = re.sub(r"\d{1,2}/\d{1,2}", "", seg_stripped)
        seg_stripped = re.sub(r"\d{1,2}日[〜～からーから]{1,3}\d{1,2}日", "", seg_stripped)
        day_matches = re.findall(r"(?<!\d)(\d{1,2})日", seg_stripped)
        if day_matches:
            month_m = re.search(r"(\d{1,2})月", seg)
            mhint = int(month_m.group(1)) if month_m else None
            for day_str in day_matches:
                found_dates += _resolve_dates([int(day_str)], mhint, dates)

        # 重複除去・ソート
        found_dates = list(dict.fromkeys(found_dates))

        if not found_dates and not out_of_range_warned:
            warnings.append(f"「{seg}」から日付を読み取れませんでした")
            continue

        for d in found_dates:
            label = _LABEL_MAP.get((shift_code, is_fixed), shift_code)
            parsed.append({
                "staff_id":   staff_id,
                "staff_name": staff_name,
                "date":       d,
                "shift":      shift_code,
                "is_fixed":   is_fixed,
                "label":      label,
            })

    return parsed, warnings, errors


# ──────────────────────────────────────────────────────────────
# スタッフ1人分テキストパーサー（Excel・テキスト共用）
# ──────────────────────────────────────────────────────────────

def _parse_per_staff_text(
    text: str,
    dates: List[datetime.date],
) -> List[Dict]:
    """
    スタッフ1人分の希望テキストを解析して [{date, shift, is_fixed}, ...] を返す。

    対応形式:
      - M/D 形式の日付: 7/21、8/4 など
      - D日 形式: 27日、25日 など
      - 数字単独（文脈依存）: 準・深の直前、または ・ の後続
      - 日付範囲: 7/21-24、11日〜14日 など
      - ・ 区切り: 複数日付をまとめて1つのシフトに / 各日付に別シフトを割り当て
      - キーワード: 準(夜勤)/深(夜)/認定日/主任会/日勤/夜勤/休み/有休/研修/委員会 等
    """
    if not text:
        return []
    text = str(text).strip()
    if text.lower() in ("nan", "none", ""):
        return []

    date_by_md: Dict[Tuple[int, int], datetime.date] = {
        (d.month, d.day): d for d in dates
    }
    months: List[int] = list(dict.fromkeys(d.month for d in dates))

    def find_shift(s: str) -> Tuple[Optional[str], bool]:
        for kw, code, fixed in _SHIFT_KW_TABLE:
            if kw in s:
                return code, fixed
        return None, False

    def resolve_day(
        day: int,
        mhint: Optional[int],
        strict: bool = False,
    ) -> Optional[datetime.date]:
        """
        day を期間内の日付に解決する。
        strict=True のとき、mhint 月に日付がなければ None（他の月へのフォールバックなし）。
        """
        if mhint:
            d = date_by_md.get((mhint, day))
            if d:
                return d
            if strict:
                return None   # 明示された月に存在しない → 誤解決を防ぐ
        for m in months:
            d = date_by_md.get((m, day))
            if d:
                return d
        return None

    results: List[Dict] = []
    last_month: Optional[int] = None

    for clause in re.split(r'[、，,\n]+', text):
        clause = clause.strip()
        if not clause:
            continue
        # 解析困難なものはスキップ
        if re.search(r'専従|以外|月[中～]|か休', clause):
            continue

        # ── 日付範囲の展開 ──
        range_dates: List[datetime.date] = []
        had_range_match = False   # 範囲パターンが一致した（期間外でも true）

        # "M/D-M/D" 形式（最優先）: "7/27-7/30", "8/11～8/14"
        for m in re.finditer(
                r'(\d{1,2})/(\d{1,2})[〜～\-－ー](\d{1,2})/(\d{1,2})', clause):
            had_range_match = True
            m1, d1, m2, d2 = (int(m.group(1)), int(m.group(2)),
                               int(m.group(3)), int(m.group(4)))
            last_month = m2
            for d in dates:
                if ((d.month, d.day) >= (m1, d1)
                        and (d.month, d.day) <= (m2, d2)
                        and d not in range_dates):
                    range_dates.append(d)

        # "M/D-D" 形式: "7/21-24日勤", "7/27-30"
        if not had_range_match:
            for m in re.finditer(
                    r'(\d{1,2})/(\d{1,2})[〜～\-－ー](\d{1,2})(?!/)', clause):
                had_range_match = True
                month, d1, d2 = int(m.group(1)), int(m.group(2)), int(m.group(3))
                last_month = month
                for day in range(d1, min(d2, 31) + 1):
                    d = date_by_md.get((month, day))
                    if d:
                        range_dates.append(d)

        # "D日〜D日" 形式: "11日〜14日まで"
        if not had_range_match:
            for m in re.finditer(
                    r'(\d{1,2})日[〜～\-ー－から]{1,3}(\d{1,2})日', clause):
                had_range_match = True
                d1, d2 = int(m.group(1)), int(m.group(2))
                for day in range(d1, min(d2, 31) + 1):
                    d = resolve_day(day, last_month)
                    if d and d not in range_dates:
                        range_dates.append(d)

        # "D-D" 形式（月コンテキスト依存）: "27-30", "22-25" など
        if not had_range_match:
            for m in re.finditer(
                    r'(?<![/\d])(\d{1,2})[〜～\-－ー](\d{1,2})(?![/日\d])', clause):
                d1, d2 = int(m.group(1)), int(m.group(2))
                if 1 <= d1 <= 31 and d1 < d2:
                    had_range_match = True
                    for day in range(d1, min(d2, 31) + 1):
                        d = resolve_day(day, last_month)
                        if d and d not in range_dates:
                            range_dates.append(d)

        # 範囲パターンが一致した場合: シフトを付与してトークナイザをスキップ
        # （期間外で range_dates が空の場合も continue して誤解決を防ぐ）
        if had_range_match:
            code, is_fixed = find_shift(clause)
            if code and range_dates:
                for d in range_dates:
                    results.append({"date": d, "shift": code, "is_fixed": is_fixed})
            if range_dates:
                last_month = range_dates[-1].month
            continue

        # ── トークン化（位置, 種類, 値, 月） ──
        tokens: List[Tuple] = []

        # M/D 形式の日付
        # clause_month: このclause内で明示された月（bare数字解決の厳格化に使う）
        clause_month: Optional[int] = None
        for m in re.finditer(r'(\d{1,2})/(\d{1,2})', clause):
            month, day = int(m.group(1)), int(m.group(2))
            clause_month = month           # 明示月を記録（期間外でも更新）
            d = date_by_md.get((month, day))
            if d:
                tokens.append((m.start(), 'date', d, month))

        md_spans = [
            (m.start(), m.end())
            for m in re.finditer(r'\d{1,2}/\d{1,2}', clause)
        ]

        def in_md(pos: int) -> bool:
            return any(s <= pos < e for s, e in md_spans)

        # clause 内に M/D があれば strict=True でその月のみ参照
        # なければ last_month ヒント＋全月フォールバック
        _strict = clause_month is not None
        _mhint  = clause_month if clause_month else last_month

        # D日 形式の日付
        for m in re.finditer(r'(?<![/\d])(\d{1,2})日', clause):
            if not in_md(m.start()):
                day = int(m.group(1))
                d = resolve_day(day, _mhint, strict=_strict)
                if d:
                    tokens.append((m.start(), 'date', d, d.month))

        # シフトキーワードの直前にある数字
        # 全角コロン「：」や半角コロン「:」も区切りとして許容（例「18：認定日」）
        for kw, code, fixed in _SHIFT_KW_TABLE:
            for m in re.finditer(
                    r'(?<![/\d])(\d{1,2})[\s：:]*' + re.escape(kw), clause):
                if not in_md(m.start()):
                    already = any(
                        t[1] == 'date' and abs(t[0] - m.start()) < 4
                        for t in tokens
                    )
                    if not already:
                        day = int(m.group(1))
                        d = resolve_day(day, _mhint, strict=_strict)
                        if d:
                            tokens.append((m.start(), 'date', d, d.month))

        # ・ の前の単独数字（例「14・19委員会」の「14」）
        for m in re.finditer(r'(?<![/\d])(\d{1,2})・', clause):
            if not in_md(m.start()):
                already = any(
                    t[1] == 'date' and abs(t[0] - m.start()) < 2
                    for t in tokens
                )
                if not already:
                    day = int(m.group(1))
                    d = resolve_day(day, _mhint, strict=_strict)
                    if d:
                        tokens.append((m.start(), 'date', d, d.month))

        # ・ の後の単独数字（例「8/4・18：認定日」の「18」）
        for m in re.finditer(r'・(\d{1,2})(?![\d/])', clause):
            digit_pos = m.start() + 1
            if not in_md(digit_pos):
                already = any(
                    t[1] == 'date' and abs(t[0] - digit_pos) < 2
                    for t in tokens
                )
                if not already:
                    day = int(m.group(1))
                    d = resolve_day(day, _mhint, strict=_strict)
                    if d:
                        tokens.append((digit_pos, 'date', d, d.month))

        # シフトキーワード（重複しないように）
        used: List[Tuple[int, int]] = []
        for kw, code, fixed in _SHIFT_KW_TABLE:
            for m in re.finditer(re.escape(kw), clause):
                if not any(s <= m.start() < e for s, e in used):
                    tokens.append((m.start(), 'shift', (code, fixed), None))
                    used.append((m.start(), m.end()))

        if not tokens:
            continue

        tokens.sort(key=lambda t: t[0])

        # ── グルーピング: 日付群 → 次のシフト ──
        pending: List[Tuple[datetime.date, Optional[int]]] = []
        for tok in tokens:
            if tok[1] == 'date':
                pending.append((tok[2], tok[3]))
            elif tok[1] == 'shift' and pending:
                code, is_fixed = tok[2]
                for d, m in pending:
                    results.append({"date": d, "shift": code, "is_fixed": is_fixed})
                    if m:
                        last_month = m
                pending = []

        # 月コンテキスト更新
        for tok in tokens:
            if tok[1] == 'date' and tok[3]:
                last_month = tok[3]

    # 重複除去（同一日付は後勝ち）
    seen: Dict[datetime.date, Dict] = {}
    for r in results:
        seen[r['date']] = r
    return list(seen.values())


# ──────────────────────────────────────────────────────────────
# Excel アップロードパーサー
# ──────────────────────────────────────────────────────────────

def _parse_date_header(val, dates: List[datetime.date]) -> Optional[datetime.date]:
    """列ヘッダー値を期間内の日付に解決する。"""
    dates_set = set(dates)
    if isinstance(val, datetime.datetime):
        d = val.date()
        return d if d in dates_set else None
    if isinstance(val, datetime.date):
        return val if val in dates_set else None
    s = str(val).strip()
    # "M/D" または "M/D(曜)" 形式
    m = re.match(r"^(\d{1,2})/(\d{1,2})", s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        for d in dates:
            if d.month == month and d.day == day:
                return d
        return None
    # "M月D日" 形式
    m = re.match(r"^(\d{1,2})月(\d{1,2})日", s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        for d in dates:
            if d.month == month and d.day == day:
                return d
        return None
    # "D日" 形式（月省略）
    m = re.match(r"^(\d{1,2})日$", s)
    if m:
        day = int(m.group(1))
        for d in dates:
            if d.day == day:
                return d
        return None
    # 数値のみ（日のみ）
    m = re.match(r"^(\d{1,2})$", s)
    if m:
        day = int(m.group(1))
        matched = [d for d in dates if d.day == day]
        return matched[0] if len(matched) == 1 else None
    return None


def _parse_cell_value(val) -> Optional[Tuple[str, bool]]:
    """セル値を (shift_code, is_fixed) に変換。空/無効は None。"""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s in ("－", "-", "−", "―", "nan", "None"):
        return None
    for kw, code, fixed in _SHIFT_KW_TABLE:
        if kw in s:
            return code, fixed
    return None


def parse_excel_requests(
    excel_bytes,
    staff_df: pd.DataFrame,
    dates: List[datetime.date],
) -> Tuple[List[Dict], List[str], List[str]]:
    """
    アップロードされた Excel (.xlsx) を解析して希望リストを返す。

    2種類のフォーマットを自動判別:

    ① 氏名＋テキスト形式（推奨）
       A列: 氏名、B列: 希望テキスト（自由記述）
       例: 「7/21認定日、7/28委員会、8/4・18休み」

    ② 氏名＋日付列形式
       A列: 氏名、以降: 日付ヘッダー列（6/21 など）
       各セル: 日勤希望・夜勤確定・休み など

    Returns:
        parsed, warnings, errors
    """
    parsed: List[Dict] = []
    warnings: List[str] = []
    errors: List[str] = []

    try:
        df = pd.read_excel(excel_bytes, header=None, dtype=str)
    except Exception as e:
        errors.append(f"Excel の読み込みに失敗しました: {e}")
        return parsed, warnings, errors

    if df.empty or df.shape[1] < 2:
        errors.append("Excel が空か、列が足りません（氏名列 + 希望列が必要）。")
        return parsed, warnings, errors

    # ── スタッフ名 → ID マップ ──
    name_to_sid: Dict[str, int] = {}
    for _, row in staff_df.iterrows():
        name: str = str(row["name"])
        sid: int = row["id"]
        name_to_sid[name] = sid
        name_to_sid[name.replace(" ", "").replace("　", "")] = sid
        parts = name.split()
        if parts:
            name_to_sid[parts[0]] = sid

    def find_staff_id(raw: str) -> Optional[int]:
        raw = raw.strip()
        sid = name_to_sid.get(raw)
        if sid:
            return sid
        raw_ns = raw.replace(" ", "").replace("　", "")
        sid = name_to_sid.get(raw_ns)
        if sid:
            return sid
        for key, s in name_to_sid.items():
            if key and (key in raw or raw in key):
                return s
        return None

    # ── フォーマット判別 ──
    # 1行目の値が日付として解釈できるかで判断
    header_row = df.iloc[0].tolist()
    date_header_count = sum(
        1 for v in header_row[1:]
        if _parse_date_header(v, dates) is not None
    )
    is_text_format = (df.shape[1] <= 5) or (date_header_count < max(1, df.shape[1] // 2))

    if is_text_format:
        # ── ① 氏名＋テキスト形式 ──
        # タイトル行スキップ判定（1行目がスタッフ名でなければスキップ）
        start_row = 0
        first_val = str(df.iloc[0, 0]).strip()
        if (not first_val
                or first_val.lower() in ("nan", "none")
                or find_staff_id(first_val) is None):
            start_row = 1

        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            raw_name = str(row.iloc[0]).strip()
            if not raw_name or raw_name.lower() in ("nan", "none"):
                continue

            staff_id = find_staff_id(raw_name)
            if staff_id is None:
                continue  # スタッフ名として認識できない行は無言スキップ

            staff_name = staff_df.loc[
                staff_df["id"] == staff_id, "name"
            ].values[0]

            # 2列目以降の非空値をテキストとして結合
            text_parts = []
            for ci in range(1, len(row)):
                val = str(row.iloc[ci]).strip()
                if val and val.lower() not in ("nan", "none"):
                    text_parts.append(val)
            text = "、".join(text_parts)
            if not text:
                continue

            for r in _parse_per_staff_text(text, dates):
                parsed.append({
                    "staff_id":   staff_id,
                    "staff_name": staff_name,
                    "date":       r["date"],
                    "shift":      r["shift"],
                    "is_fixed":   r["is_fixed"],
                    "label":      _LABEL_MAP.get((r["shift"], r["is_fixed"]), r["shift"]),
                })

    else:
        # ── ② 氏名＋日付列形式 ──
        col_to_date: Dict[int, datetime.date] = {}
        for ci, val in enumerate(header_row[1:], 1):
            d = _parse_date_header(val, dates)
            if d is not None:
                col_to_date[ci] = d

        if not col_to_date:
            errors.append(
                "日付として認識できる列ヘッダーがありませんでした。"
                "「6/21」「6月21日」「21」などの形式で記載してください。"
            )
            return parsed, warnings, errors

        for idx in range(1, len(df)):
            row = df.iloc[idx]
            raw_name = str(row.iloc[0]).strip()
            if not raw_name or raw_name.lower() in ("nan", "none"):
                continue
            staff_id = find_staff_id(raw_name)
            if staff_id is None:
                warnings.append(f"スタッフ「{raw_name}」が見つかりませんでした（{idx+1}行目）")
                continue
            staff_name = staff_df.loc[
                staff_df["id"] == staff_id, "name"
            ].values[0]

            for ci, date in col_to_date.items():
                result = _parse_cell_value(row.iloc[ci] if ci < len(row) else None)
                if result is None:
                    continue
                shift_code, is_fixed = result
                parsed.append({
                    "staff_id":   staff_id,
                    "staff_name": staff_name,
                    "date":       date,
                    "shift":      shift_code,
                    "is_fixed":   is_fixed,
                    "label":      _LABEL_MAP.get((shift_code, is_fixed), shift_code),
                })

    # ── 期間外日付が多い場合の警告 ──────────────────────────────
    # テキスト中の M/D 日付を全スキャンして期間内 vs 期間外の比率を確認
    if not errors:
        date_by_md_set = {(d.month, d.day) for d in dates}
        in_period_cnt  = 0
        out_period_cnt = 0
        text_col_range = range(1, df.shape[1]) if is_text_format else range(0, 0)
        scan_df = df.iloc[start_row:] if is_text_format else df.iloc[1:]
        for _, row in scan_df.iterrows():
            cell_text = " ".join(
                str(row.iloc[ci])
                for ci in (text_col_range if is_text_format else [])
                if str(row.iloc[ci]).strip().lower() not in ("nan", "none", "")
            )
            for mm in re.finditer(r'(\d{1,2})/(\d{1,2})', cell_text):
                key = (int(mm.group(1)), int(mm.group(2)))
                if key in date_by_md_set:
                    in_period_cnt += 1
                else:
                    out_period_cnt += 1
        if out_period_cnt > 0 and out_period_cnt > in_period_cnt:
            period_str = f"{dates[0].month}/{dates[0].day}〜{dates[-1].month}/{dates[-1].day}"
            warnings.append(
                f"⚠️ Excelに含まれる日付（M/D形式）の多くが対象期間外でした"
                f"（期間内: {in_period_cnt}件、期間外: {out_period_cnt}件）。"
                f"サイドバーの「開始月」が正しいか確認してください（現在の期間: {period_str}）。"
            )

    return parsed, warnings, errors


def make_template_excel(
    staff_df: pd.DataFrame,
    dates: List[datetime.date],
) -> bytes:
    """
    希望入力用テンプレート Excel (.xlsx) を生成して bytes で返す。

    フォーマット（アップロードされたファイルと同じ形式）:
      A列: 氏名（スタッフ名を事前記入）
      B列: 希望テキスト（自由記入）
      1行目: タイトル行
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "勤務希望"

    THIN   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL = PatternFill("solid", fgColor="37474F")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=9)

    # ── 期間タイトル ──
    months = list(dict.fromkeys(d.month for d in dates))
    period_str = "・".join(f"{m}月" for m in months) + "の勤務希望"
    c = ws.cell(1, 1, period_str)
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 18

    # ── ヘッダー行（2行目） ──
    for col_idx, label in enumerate(["氏名", "希望テキスト（自由記入）"], 1):
        c = ws.cell(2, col_idx, label)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 70

    # ── スタッフ行（3行目〜） ──
    for ri, (_, row) in enumerate(staff_df.iterrows(), 3):
        c = ws.cell(ri, 1, row["name"])
        c.border = BORDER
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="center")
        c2 = ws.cell(ri, 2, "")
        c2.border = BORDER
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = 20

    # ── 記入例（スタッフ行の後に空行を1つ空けて） ──
    ex_row = len(staff_df) + 4
    ws.cell(ex_row, 1, "【記入例】").font = Font(bold=True, size=8, color="888888")
    examples = (
        "7/24準・25深、27日研修、8/7委員会、8・9休み、14・19委員会（/イ）、"
        "21日準、22日深夜、24日休み"
    )
    c = ws.cell(ex_row, 2, examples)
    c.font = Font(size=8, color="888888", italic=True)

    # ── キーワード一覧 ──
    kw_row = ex_row + 1
    ws.cell(kw_row, 1, "【キーワード】").font = Font(bold=True, size=8, color="888888")
    kw_text = (
        "準／準夜勤=夜勤入り　深／深夜勤=夜勤明け　"
        "日勤希望　夜勤希望　夜勤確定　休み希望　有休申請　"
        "研修　認定日　委員会　主任会　遅出"
    )
    c = ws.cell(kw_row, 2, kw_text)
    c.font = Font(size=8, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# カレンダー描画
# ──────────────────────────────────────────────────────────────

def _badge_html(short: str) -> str:
    """カレンダーセル内バッジ（値が「－」のときは空文字を返す）。"""
    if short == "－" or short not in _BADGE_COLOR:
        return ""
    bg, fg = _BADGE_COLOR[short]
    return (
        f"<div style='text-align:center;background:{bg};color:{fg};"
        f"border-radius:6px;font-size:13px;font-weight:bold;"
        f"padding:3px 4px;margin:2px 0;letter-spacing:0.02em'>{short}</div>"
    )


def _badge_inline_html(short: str) -> str:
    """サマリーパネル用インラインバッジ（大きめ）。"""
    if short not in _BADGE_COLOR:
        return short
    bg, fg = _BADGE_COLOR[short]
    return (
        f"<span style='background:{bg};color:{fg};"
        f"border-radius:6px;font-size:15px;font-weight:bold;"
        f"padding:3px 14px;display:inline-block'>{short}</span>"
    )


def _render_staff_summary(
    selected_name: str,
    selected_sid: int,
    new_rows: list,
    jp_hols,
) -> None:
    """
    選択中スタッフの入力済み希望を見やすく一覧表示。
    削除ボタンで個別に解除できる。
    """
    WDAY = "月火水木金土日"

    if not new_rows:
        st.info(f"📭 {selected_name} の希望はまだ登録されていません。")
        return

    fixed_cnt  = sum(1 for r in new_rows if r["is_fixed"])
    soft_cnt   = len(new_rows) - fixed_cnt
    st.markdown(
        f"<div style='font-size:14px;font-weight:bold;margin:6px 0 10px'>"
        f"📋 入力済み：{len(new_rows)}件"
        f"　<span style='color:#b71c1c;font-size:12px'>🔒確定 {fixed_cnt}件</span>"
        f"　<span style='color:#1565c0;font-size:12px'>💭希望 {soft_cnt}件</span>"
        f"</div>",
        unsafe_allow_html=True,
    )

    to_delete = None
    # ヘッダー行
    h0, h1, h2, h3, h4 = st.columns([3, 2, 2, 2, 1])
    for h, label in zip([h0, h1, h2, h3], ["日付", "希望", "種別", ""]):
        h.markdown(
            f"<div style='font-size:11px;color:#888;border-bottom:1px solid #ddd;"
            f"padding-bottom:2px;margin-bottom:4px'>{label}</div>",
            unsafe_allow_html=True,
        )

    for row in sorted(new_rows, key=lambda r: r["date"]):
        d        = row["date"]
        shift    = row["shift"]
        is_fixed = row["is_fixed"]
        short    = _FULL_TO_SHORT.get(_code_to_option(shift, is_fixed), "?")
        wday     = WDAY[d.weekday()]
        is_hol   = d.weekday() == 6 or d in jp_hols
        is_sat   = d.weekday() == 5
        date_clr = _HOL_COLOR if is_hol else (_SAT_COLOR if is_sat else _NORM_COLOR)

        c0, c1, c2, c3, c4 = st.columns([3, 2, 2, 2, 1])
        c0.markdown(
            f"<div style='color:{date_clr};font-size:14px;font-weight:bold;"
            f"padding-top:5px'>{d.month}/{d.day}（{wday}）</div>",
            unsafe_allow_html=True,
        )
        c1.markdown(
            _badge_inline_html(short),
            unsafe_allow_html=True,
        )
        c2.markdown(
            f"<div style='font-size:13px;padding-top:5px;"
            f"color:{'#b71c1c' if is_fixed else '#1565c0'}'>"
            f"{'🔒 確定' if is_fixed else '💭 希望'}</div>",
            unsafe_allow_html=True,
        )
        if c4.button("×", key=f"sum_del_{selected_sid}_{d}", help="この希望を削除"):
            to_delete = d

    if to_delete is not None:
        wkey = f"req_cal_w_{selected_sid}_{to_delete}"
        st.session_state[wkey] = "－"
        st.rerun()


def _code_to_option(shift_code: str, is_fixed: bool) -> str:
    for label, (code, fixed) in REQUEST_TO_SHIFT.items():
        if code == shift_code and fixed == is_fixed:
            return label
    return "（なし）"


def _day_color(d: datetime.date, jp_hols) -> str:
    if d.weekday() == 6 or d in jp_hols:
        return _HOL_COLOR
    if d.weekday() == 5:
        return _SAT_COLOR
    return _NORM_COLOR


def _render_month_calendar(
    year: int,
    month: int,
    schedule_dates_set: set,
    selected_sid: int,
    jp_hols,
) -> None:
    st.markdown(
        f"<div style='font-size:15px;font-weight:bold;margin:8px 0 4px'>"
        f"{year}年{month}月</div>",
        unsafe_allow_html=True,
    )
    header_cols = st.columns(7)
    header_colors = [_NORM_COLOR] * 5 + [_SAT_COLOR, _HOL_COLOR]
    for i, (wn, hc) in enumerate(zip(WEEKDAY_NAMES, header_colors)):
        header_cols[i].markdown(
            f"<div style='text-align:center;color:{hc};font-size:12px;"
            f"font-weight:bold;border-bottom:1px solid #ddd;padding-bottom:2px'>{wn}</div>",
            unsafe_allow_html=True,
        )

    first_day = datetime.date(year, month, 1)
    last_day  = datetime.date(year, month, cal_module.monthrange(year, month)[1])
    start = first_day - datetime.timedelta(days=first_day.weekday())
    end   = last_day  + datetime.timedelta(days=6 - last_day.weekday())

    current = start
    while current <= end:
        week_cols = st.columns(7)
        for i in range(7):
            d = current + datetime.timedelta(days=i)
            with week_cols[i]:
                if d.month != month:
                    # 当月以外は空白
                    st.markdown("<div style='min-height:72px'></div>",
                                unsafe_allow_html=True)
                elif d in schedule_dates_set:
                    wkey = f"req_cal_w_{selected_sid}_{d}"
                    chosen = st.session_state.get(wkey, "－")
                    color = _day_color(d, jp_hols)

                    # ① 大きめバッジ（設定済みの場合のみ）
                    badge = _badge_html(chosen)
                    if badge:
                        st.markdown(badge, unsafe_allow_html=True)
                    else:
                        st.markdown("<div style='min-height:4px'></div>",
                                    unsafe_allow_html=True)

                    # ② 日付
                    st.markdown(
                        f"<div style='text-align:center;color:{color};"
                        f"font-size:12px;font-weight:bold;margin-bottom:1px'>"
                        f"{d.day}</div>",
                        unsafe_allow_html=True,
                    )

                    # ③ セレクトボックス（変更用）
                    st.selectbox(
                        f"sel_{selected_sid}_{d}",
                        SHORT_OPTIONS,
                        index=SHORT_OPTIONS.index(chosen)
                              if chosen in SHORT_OPTIONS else 0,
                        key=wkey,
                        label_visibility="collapsed",
                    )
                else:
                    # 対象期間外の日付
                    st.markdown(
                        f"<div style='text-align:center;color:{_GRAY_COLOR};"
                        f"font-size:12px;min-height:72px;padding-top:4px'>{d.day}</div>",
                        unsafe_allow_html=True,
                    )
        current += datetime.timedelta(days=7)


# ──────────────────────────────────────────────────────────────
# メイン関数
# ──────────────────────────────────────────────────────────────

def render_request_calendar(
    staff_df: pd.DataFrame,
    dates: List[datetime.date],
) -> None:
    st.subheader("勤務希望カレンダー")

    sid_to_name = {row["id"]: row["name"] for _, row in staff_df.iterrows()}
    name_to_sid = {v: k for k, v in sid_to_name.items()}
    staff_names = staff_df["name"].tolist()
    existing: pd.DataFrame = st.session_state.requests_df.copy()

    # ────────────────────────────────
    # テキスト一括入力セクション
    # ────────────────────────────────
    with st.expander("💬 テキストで一括入力", expanded=True):
        st.caption(
            "例: 「秋田は7月24日は夜勤確定、7日は日勤、8日9日は休み」\n"
            "キーワード: 日勤／夜勤／休み／有休／研修／遅出　　「確定」を付けるとハード条件"
        )
        input_text = st.text_area(
            "希望テキストを入力",
            height=80,
            placeholder="例: 小島は6月25日は夜勤希望、7月3日4日は有休",
            label_visibility="collapsed",
            key="req_text_input",
        )
        if st.button("解析・適用", key="req_text_parse", type="primary"):
            if input_text.strip():
                parsed, warnings, errors = parse_request_text(
                    input_text, staff_df, dates
                )
                if errors:
                    for e in errors:
                        st.error(e)
                elif not parsed:
                    st.warning("日付とシフトを読み取れませんでした。")
                else:
                    # requests_df を更新（対象スタッフの該当日のみ上書き）
                    updated = existing.copy()
                    for r in parsed:
                        mask = (
                            (updated["staff_id"] == r["staff_id"]) &
                            (updated["date"] == r["date"])
                        )
                        updated = updated[~mask]
                    new_df = pd.DataFrame(
                        [{"staff_id": r["staff_id"], "date": r["date"],
                          "shift": r["shift"], "is_fixed": r["is_fixed"]}
                         for r in parsed],
                        columns=["staff_id", "date", "shift", "is_fixed"],
                    )
                    st.session_state.requests_df = pd.concat(
                        [updated, new_df], ignore_index=True
                    )
                    save_requests(st.session_state.requests_df)
                    # カレンダーを再初期化させる
                    st.session_state["_req_cal_prev_sid"] = None
                    for w in warnings:
                        st.warning(w)
                    # 結果プレビュー
                    st.success(f"✅ {len(parsed)} 件の希望を適用しました")
                    preview = pd.DataFrame([
                        {"氏名": r["staff_name"],
                         "日付": r["date"].strftime("%m/%d"),
                         "希望": r["label"],
                         "種別": "確定" if r["is_fixed"] else "希望"}
                        for r in parsed
                    ])
                    st.dataframe(preview, hide_index=True, use_container_width=True)
                    st.rerun()
            else:
                st.info("テキストを入力してください。")

    # ────────────────────────────────
    # Excel アップロードセクション
    # ────────────────────────────────
    with st.expander("📊 Excelで一括入力", expanded=False):
        st.caption(
            "スタッフ名を1列目、日付を1行目に並べた Excel ファイルをアップロードすると"
            "希望を一括登録できます。テンプレートをダウンロードして使ってください。"
        )
        # テンプレートダウンロード
        tmpl_bytes = make_template_excel(staff_df, dates)
        period_label = (
            f"{dates[0].strftime('%Y%m%d')}-{dates[-1].strftime('%Y%m%d')}"
        )
        st.download_button(
            "📥 テンプレートをダウンロード",
            data=tmpl_bytes,
            file_name=f"希望テンプレート_{period_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="req_tmpl_dl",
        )

        uploaded = st.file_uploader(
            "Excelファイルを選択（.xlsx）",
            type=["xlsx"],
            key="req_excel_upload",
            label_visibility="collapsed",
        )
        if uploaded is not None:
            parsed_xl, warn_xl, err_xl = parse_excel_requests(uploaded, staff_df, dates)
            if err_xl:
                for e in err_xl:
                    st.error(e)
            elif not parsed_xl:
                st.warning("Excel から希望を読み取れませんでした。フォーマットを確認してください。")
                for w in warn_xl:
                    st.warning(w)
            else:
                for w in warn_xl:
                    st.warning(w)
                # プレビュー
                preview_xl = pd.DataFrame([
                    {"氏名": r["staff_name"],
                     "日付": r["date"].strftime("%m/%d"),
                     "希望": r["label"],
                     "種別": "確定" if r["is_fixed"] else "希望"}
                    for r in parsed_xl
                ])
                st.info(f"✅ {len(parsed_xl)} 件の希望が読み取られました。確認して「適用」してください。")
                st.dataframe(preview_xl, hide_index=True, use_container_width=True, height=200)

                if st.button("Excel 希望を適用", key="req_excel_apply", type="primary"):
                    updated_xl = existing.copy()
                    for r in parsed_xl:
                        mask = (
                            (updated_xl["staff_id"] == r["staff_id"]) &
                            (updated_xl["date"] == r["date"])
                        )
                        updated_xl = updated_xl[~mask]
                    new_df_xl = pd.DataFrame(
                        [{"staff_id": r["staff_id"], "date": r["date"],
                          "shift": r["shift"], "is_fixed": r["is_fixed"]}
                         for r in parsed_xl],
                        columns=["staff_id", "date", "shift", "is_fixed"],
                    )
                    st.session_state.requests_df = pd.concat(
                        [updated_xl, new_df_xl], ignore_index=True
                    )
                    save_requests(st.session_state.requests_df)
                    st.session_state["_req_cal_prev_sid"] = None
                    st.success(f"{len(parsed_xl)} 件の希望を登録しました。")
                    st.rerun()

    st.divider()

    # ────────────────────────────────
    # 個人カレンダー選択・表示
    # ────────────────────────────────
    col_sel, col_info = st.columns([2, 3])
    with col_sel:
        selected_name = st.selectbox("スタッフを選択", staff_names, key="req_cal_staff")
    selected_sid = name_to_sid[selected_name]

    with col_info:
        if len(existing) > 0:
            st.caption(f"全スタッフ 希望入力数: {len(existing)} 件")

    # 凡例
    legend = "　".join(
        f'<span style="background:{_BADGE_COLOR[s][0]};color:{_BADGE_COLOR[s][1]};'
        f'border-radius:3px;padding:1px 5px;font-size:11px;font-weight:bold">{s}</span>'
        f'={f}'
        for s, f in zip(SHORT_OPTIONS[1:], REQUEST_OPTIONS[1:])
    )
    st.markdown(legend, unsafe_allow_html=True)

    # スタッフ切り替え または 表示期間変更 時に widget キーを requests_df から再初期化
    _dates_key = f"{dates[0]}_{dates[-1]}"
    if (st.session_state.get("_req_cal_prev_sid") != selected_sid
            or st.session_state.get("_req_cal_prev_dates") != _dates_key):
        staff_reqs = existing[existing["staff_id"] == selected_sid]
        existing_prefs: dict = {
            row["date"]: _code_to_option(row["shift"], row["is_fixed"])
            for _, row in staff_reqs.iterrows()
        }
        for d in dates:
            wkey = f"req_cal_w_{selected_sid}_{d}"
            short_val = _FULL_TO_SHORT.get(existing_prefs.get(d, "（なし）"), "－")
            if wkey in st.session_state:
                del st.session_state[wkey]
            st.session_state[wkey] = short_val
        st.session_state["_req_cal_prev_sid"] = selected_sid
        st.session_state["_req_cal_prev_dates"] = _dates_key

    # カレンダー描画
    schedule_dates_set = set(dates)
    months = list(dict.fromkeys((d.year, d.month) for d in dates))
    jp_hols = holidays_lib.Japan(years={d.year for d in dates})

    # ── 2ペイン レイアウト: カレンダー(左) | 入力済み一覧(右) ──
    col_cal, col_summary = st.columns([3, 2])

    with col_cal:
        if len(months) == 2:
            # 2ヶ月を上下に並べる
            _render_month_calendar(months[0][0], months[0][1],
                                   schedule_dates_set, selected_sid, jp_hols)
            st.markdown("<div style='margin:8px 0'></div>", unsafe_allow_html=True)
            _render_month_calendar(months[1][0], months[1][1],
                                   schedule_dates_set, selected_sid, jp_hols)
        else:
            for ym in months:
                _render_month_calendar(ym[0], ym[1],
                                       schedule_dates_set, selected_sid, jp_hols)

    # widget キーから requests_df を更新
    new_rows: list = []
    for d in dates:
        wkey = f"req_cal_w_{selected_sid}_{d}"
        short = st.session_state.get(wkey, "－")
        if short != "－":
            full = _SHORT_TO_FULL.get(short, "（なし）")
            if full in REQUEST_TO_SHIFT:
                shift_code, is_fixed = REQUEST_TO_SHIFT[full]
                new_rows.append({
                    "staff_id": selected_sid,
                    "date": d,
                    "shift": shift_code,
                    "is_fixed": is_fixed,
                })

    # 当該スタッフの「期間外」レコードは保持する（別の期間の希望を消さないため）
    dates_set = set(dates)
    other = existing[existing["staff_id"] != selected_sid]
    selected_outside = existing[
        (existing["staff_id"] == selected_sid) & ~existing["date"].isin(dates_set)
    ]
    merged = pd.concat(
        [other, selected_outside,
         pd.DataFrame(new_rows, columns=["staff_id", "date", "shift", "is_fixed"])]
        if new_rows else [other, selected_outside],
        ignore_index=True,
    )
    if not merged.reset_index(drop=True).equals(
            st.session_state.requests_df.reset_index(drop=True)):
        st.session_state.requests_df = merged

    # ── 右ペイン: 入力済み一覧 ──
    with col_summary:
        st.markdown(
            f"<div style='font-size:15px;font-weight:bold;color:#333;"
            f"margin-bottom:6px'>👤 {selected_name}</div>",
            unsafe_allow_html=True,
        )
        _render_staff_summary(selected_name, selected_sid, new_rows, jp_hols)

    # 全スタッフ希望一覧
    if not st.session_state.requests_df.empty:
        with st.expander("全スタッフの希望一覧を表示"):
            view = st.session_state.requests_df.copy()
            view["氏名"] = view["staff_id"].map(sid_to_name)
            view["日付"] = view["date"].astype(str)
            view["希望"] = view.apply(
                lambda r: _code_to_option(r["shift"], r["is_fixed"]), axis=1)
            view["種別"] = view["is_fixed"].map({True: "確定", False: "希望"})
            st.dataframe(
                view[["氏名", "日付", "希望", "種別"]].sort_values(["氏名", "日付"]),
                use_container_width=True, hide_index=True,
            )
        if st.button("全希望をクリア", key="clear_all_requests"):
            st.session_state.requests_df = pd.DataFrame(
                columns=["staff_id", "date", "shift", "is_fixed"])
            save_requests(st.session_state.requests_df)
            for d in dates:
                wkey = f"req_cal_w_{selected_sid}_{d}"
                if wkey in st.session_state:
                    del st.session_state[wkey]
            st.rerun()
